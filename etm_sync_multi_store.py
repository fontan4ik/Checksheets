# -*- coding: utf-8 -*-
import argparse
import csv
import io
import ipaddress
import json
import logging
import os
import posixpath
import re
import socket
import subprocess
import sys
import time
import zipfile
from dataclasses import dataclass
from datetime import datetime, timedelta, timezone
from ftplib import FTP, FTP_TLS, error_perm
from pathlib import Path
from typing import Dict, Iterable, List, Optional
from xml.etree import ElementTree as ET

import gspread
import requests

import config
import gsheets_utils
from network_bypass import SourceAddressAdapter


LOG_PATH = os.path.join(os.path.dirname(__file__), "logs", "etm_sync_multi.log")
os.makedirs(os.path.dirname(LOG_PATH), exist_ok=True)

ETM_TR_SCHEMA = {
    "etm_code": "CODES",
    "stock_nsb": "ETM NSB",
    "stock_smr": "ETM SMR",
}
FERON_ETM_MAPPING_PATH = Path(__file__).resolve().parent / "СОПОСТАВЛЕННОЕ ЭТМ .xlsx"
STREAM_SUPPS_SHEET_NAME = getattr(config, "STREAM_SUPPS_SHEET_NAME", "StreamSupps")
SHEETS_UPDATE_RETRIES = 5
SHEETS_UPDATE_RETRY_DELAY = 2.0

FTP_HOST = os.getenv("ETM_FTP_HOST", getattr(config, "ETM_FTP_HOST", "edi.etm.ru"))
FTP_PORT = int(os.getenv("ETM_FTP_PORT", getattr(config, "ETM_FTP_PORT", "21")))
FTP_USER = os.getenv("ETM_FTP_USER", getattr(config, "ETM_FTP_USER", "u_energoservis"))
FTP_PASSWORD = os.getenv("ETM_FTP_PASSWORD", getattr(config, "ETM_FTP_PASSWORD", ""))
FTP_TIMEOUT = int(os.getenv("ETM_FTP_TIMEOUT", "300"))
FTP_WAREHOUSE_RETRIES = max(1, int(os.getenv("ETM_FTP_WAREHOUSE_RETRIES", "3")))
FTP_WAREHOUSE_RETRY_DELAY = max(0.0, float(os.getenv("ETM_FTP_WAREHOUSE_RETRY_DELAY", "2")))
FTP_TLS_MODE = os.getenv("ETM_FTP_TLS", getattr(config, "ETM_FTP_TLS", "disable")).strip().lower()
FTP_LOCAL_ROOT = Path(
    os.getenv(
        "ETM_FTP_LOCAL_ROOT",
        str(Path(__file__).resolve().parent / "test" / "tmp" / "etm_ftp_downloads"),
    )
)
FTP_PROCESS_MODE = os.getenv("ETM_FTP_PROCESS_MODE", "latest").strip().lower()
FTP_REQUIRE_TODAY = os.getenv("ETM_FTP_REQUIRE_TODAY", "1").strip().lower() not in {
    "0",
    "false",
    "no",
}
FTP_STATE_PATH = Path(
    os.getenv(
        "ETM_FTP_STATE_PATH",
        str(Path(__file__).resolve().parent / "test" / "tmp" / "etm_ftp_state.json"),
    )
)

WAREHOUSE_DIRS = {
    "smr": {
        "remote_dir": os.getenv("ETM_FTP_SMR_DIR", "/from_etm/13"),
        "header": os.getenv("ETM_FTP_SMR_HEADER", "stocks smr"),
        "label": "Samara",
    },
    "nsb": {
        "remote_dir": os.getenv("ETM_FTP_NSB_DIR", "/from_etm/16"),
        "header": os.getenv("ETM_FTP_NSB_HEADER", "stocks nsb"),
        "label": "Novosibirsk",
    },
}

STOCK_FIELD_NAMES = {
    "rem",
    "reminfo",
    "stock",
    "stocks",
    "qty",
    "quantity",
    "amount",
    "available",
    "free",
    "balance",
    "rest",
    "ostatok",
    "ostatki",
    "storequantrem",
    "quant",
    "количество",
    "остаток",
    "остатки",
    "доступно",
}

CODE_FIELD_NAMES = {
    "gdscode",
    "gds",
    "etmcode",
    "etm_code",
    "supplieritemcode",
    "suppliercode",
    "supplier_code",
    "code",
    "id",
    "кодэтм",
    "код_этм",
    "код",
}

ARTICLE_FIELD_NAMES = {
    "article",
    "art",
    "manufacturerarticle",
    "manufacturercod",
    "mnfarticle",
    "mnf_art",
    "vendorcode",
    "vendor_code",
    "sku",
    "артикул",
    "артикулпроизводителя",
}

MANUFACTURER_FIELD_NAMES = {
    "manufacturer",
    "manufacturername",
    "producer",
    "vendor",
    "brand",
    "производитель",
    "бренд",
}


logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s - %(levelname)s - %(message)s",
    handlers=[
        logging.FileHandler(LOG_PATH, encoding="utf-8", mode="w"),
        logging.StreamHandler(sys.stdout),
    ],
)


def enumerate_local_ipv4s():
    """Return usable local IPv4 candidates, preferring active LAN/Wi-Fi.

    Hostname resolution is not a reliable way to discover local addresses on
    macOS: with Network Extension VPN clients it may return no address at all
    (or only a tunnel address).  The ETM bypass must bind to the physical
    LAN/Wi-Fi source, just like the other local sync scripts.
    """
    ips = []
    seen = set()

    def add_many(values):
        for value in values:
            if not value or value == "127.0.0.1" or value in seen:
                continue
            seen.add(value)
            ips.append(value)

    preferred_interface = os.getenv("CHECKSHEETS_BYPASS_INTERFACE", "").strip()
    interfaces = [preferred_interface] if preferred_interface else ["en1", "en0"]

    for interface in interfaces:
        if not interface:
            continue
        try:
            result = subprocess.run(
                ["/sbin/ifconfig", interface],
                capture_output=True,
                text=True,
                check=False,
            )
        except OSError:
            continue

        output = result.stdout or ""
        if not preferred_interface and "status: active" not in output:
            continue
        add_many(re.findall("inet ([0-9]+[.][0-9]+[.][0-9]+[.][0-9]+)", output))

    # Keep hostname resolution as a fallback for non-standard environments,
    # but never let tunnel-only addresses replace an active LAN/Wi-Fi source.
    if not ips:
        host = socket.gethostname()
        try:
            add_many(socket.gethostbyname_ex(host)[2])
        except OSError:
            pass
        try:
            add_many(addr[4][0] for addr in socket.getaddrinfo(host, None, socket.AF_INET))
        except OSError:
            pass

    return ips


def choose_bypass_source():
    override_ip = os.getenv("CHECKSHEETS_BYPASS_IP", "").strip()
    if override_ip:
        return os.getenv("CHECKSHEETS_BYPASS_INTERFACE", "override").strip() or "override", override_ip

    preferred_interface = os.getenv("CHECKSHEETS_BYPASS_INTERFACE", "").strip()
    candidates = enumerate_local_ipv4s()

    def rank(ip):
        try:
            parsed = ipaddress.ip_address(ip)
        except ValueError:
            return (3, ip)
        if parsed.is_private:
            return (0, ip)
        if parsed.is_link_local:
            return (1, ip)
        return (2, ip)

    candidates.sort(key=rank)
    if candidates:
        return preferred_interface or "auto", candidates[0]

    raise RuntimeError(
        "No usable IPv4 source found for ETM bypass. "
        "Set CHECKSHEETS_BYPASS_IP explicitly."
    )


def create_etm_session():
    """Compatibility function for scripts that still use the ETM HTTP API."""
    label, source_ip = choose_bypass_source()
    session = requests.Session()
    session.trust_env = False
    interface_name = label if label not in {"auto", "override"} else None
    if interface_name is None:
        for candidate in ("en1", "en0"):
            output = subprocess.run(
                ["/sbin/ifconfig", candidate],
                capture_output=True,
                text=True,
                check=False,
            ).stdout
            if source_ip in output:
                interface_name = candidate
                break
    adapter = SourceAddressAdapter(source_ip, interface_name=interface_name)
    session.mount("http://", adapter)
    session.mount("https://", adapter)
    logging.info("ETM HTTP bypass source: %s (%s)", label, source_ip)
    return session


def login_etm(http, headers):
    """Compatibility function for etm_export_codes.py."""
    login_data = None
    session_id = None
    for attempt in range(3):
        try:
            logging.info("Logging into ETM HTTP API (Attempt %s/3)...", attempt + 1)
            params = {"log": config.ETM_LOGIN, "pwd": config.ETM_PASSWORD}
            response = http.post(
                "https://ipro.etm.ru/api/v1/user/login",
                params=params,
                headers=headers,
                timeout=30,
            )
            response.raise_for_status()
            login_data = response.json()
            session_id = login_data["data"]["session"]
            break
        except Exception as exc:
            logging.warning("Login attempt %s failed: %s", attempt + 1, exc)
            if attempt < 2:
                time.sleep(5)
    return session_id, login_data


@dataclass
class FtpFile:
    remote_path: str
    size: Optional[int] = None
    modified: Optional[str] = None


@dataclass
class StockRecord:
    gds_code: str
    article: str
    stock: int
    source: str
    manufacturer: str = ""


def normalize(s):
    if not s:
        return ""
    s = str(s).strip()
    s = re.sub(r"\.0+$", "", s)
    s = re.sub(r"\(.*\)", "", s)
    s = re.sub(r"[^A-Z0-9]", "", s.upper())
    return s


def normalize_field_name(value):
    value = str(value or "").strip().lower()
    value = value.replace("ё", "е")
    return re.sub(r"[^a-zа-я0-9]+", "", value)


def normalize_etm_code(value):
    raw = str(value or "").strip()
    if not raw:
        return ""
    raw = re.sub(r"\.0+$", "", raw)
    raw = re.sub(r"^ETM", "", raw, flags=re.IGNORECASE)
    digits = re.sub(r"\D", "", raw)
    return digits


def get_sheet_etm_code(row, column):
    if len(row) < column:
        return ""
    return normalize_etm_code(row[column - 1])


def normalize_mapping_key(value):
    raw = str(value or "").strip()
    return re.sub(r"\.0+$", "", raw)


def read_xlsx_first_sheet_rows(path):
    main_ns = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
    rel_ns = "http://schemas.openxmlformats.org/package/2006/relationships"
    office_rel_ns = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"

    with zipfile.ZipFile(path) as archive:
        shared_strings = []
        if "xl/sharedStrings.xml" in archive.namelist():
            shared_root = ET.fromstring(archive.read("xl/sharedStrings.xml"))
            for item in shared_root.findall(f"{{{main_ns}}}si"):
                shared_strings.append("".join(item.itertext()))

        workbook_root = ET.fromstring(archive.read("xl/workbook.xml"))
        first_sheet = workbook_root.find(f".//{{{main_ns}}}sheet")
        if first_sheet is None:
            return []
        relation_id = first_sheet.attrib[f"{{{office_rel_ns}}}id"]

        rels_root = ET.fromstring(archive.read("xl/_rels/workbook.xml.rels"))
        target = None
        for relation in rels_root.findall(f"{{{rel_ns}}}Relationship"):
            if relation.attrib.get("Id") == relation_id:
                target = relation.attrib.get("Target")
                break
        if not target:
            raise RuntimeError(f"Could not resolve first worksheet in {path}")

        sheet_path = posixpath.normpath(posixpath.join("xl", target))
        sheet_root = ET.fromstring(archive.read(sheet_path))
        rows = []
        for row_node in sheet_root.findall(f".//{{{main_ns}}}row"):
            values = {}
            for cell in row_node.findall(f"{{{main_ns}}}c"):
                ref = cell.attrib.get("r", "")
                column = re.sub(r"\d", "", ref)
                value_node = cell.find(f"{{{main_ns}}}v")
                inline_node = cell.find(f"{{{main_ns}}}is")
                value = ""
                if inline_node is not None:
                    value = "".join(inline_node.itertext())
                elif value_node is not None:
                    value = value_node.text or ""
                    if cell.attrib.get("t") == "s":
                        value = shared_strings[int(value)]
                values[column] = value
            rows.append(values)
        return rows


def load_feron_etm_mapping(path=FERON_ETM_MAPPING_PATH):
    if not path.exists():
        raise FileNotFoundError(f"FERON ETM mapping file not found: {path}")

    rows = read_xlsx_first_sheet_rows(path)
    mapping = {}
    conflicts = {}
    for row in rows[1:]:
        match_key = normalize_mapping_key(row.get("A"))
        etm_code = normalize_etm_code(row.get("B"))
        if not match_key or not etm_code:
            continue
        previous = mapping.get(match_key)
        if previous and previous != etm_code:
            conflicts.setdefault(match_key, {previous}).add(etm_code)
            continue
        mapping[match_key] = etm_code

    if conflicts:
        raise RuntimeError(f"Conflicting FERON ETM mappings: {len(conflicts)}")
    logging.info("Loaded %s FERON ETM mappings from %s", len(mapping), path.name)
    return mapping


def extract_tail_key(value):
    match = re.search(r"(\d+)$", str(value or "").strip())
    if not match:
        return ""
    tail = match.group(1)
    return tail if len(tail) >= 4 else ""


def build_match_candidates(value):
    normalized = normalize(value)
    return [normalized] if normalized else []


def should_use_tail_match(normalized_value):
    return bool(normalized_value) and normalized_value.isdigit()


def add_stock_to_lookup(lookup, gds_lookup, tail_lookup, article_to_gds, tail_to_gds, loose_entries, item):
    gds = normalize_etm_code(item.get("GdsCode") or item.get("gds_code"))
    raw_art = item.get("Article") or item.get("article") or ""
    art = normalize(raw_art)

    try:
        stock = int(float(str(item.get("RemInfo", item.get("stock", 0))).replace(",", ".")))
    except Exception:
        stock = 0

    if stock <= 0:
        return

    if gds:
        gds_lookup[gds] = gds_lookup.get(gds, 0) + stock
        lookup[gds] = lookup.get(gds, 0) + stock
    if art:
        lookup[art] = lookup.get(art, 0) + stock
        article_to_gds.setdefault(art, gds)

    dotted_base = str(raw_art).strip().split(".", 1)[0].strip()
    dotted_base_norm = normalize(dotted_base)
    if dotted_base_norm and dotted_base_norm != art:
        lookup[dotted_base_norm] = lookup.get(dotted_base_norm, 0) + stock
        article_to_gds.setdefault(dotted_base_norm, gds)

    tail = extract_tail_key(raw_art)
    if tail:
        tail_lookup[tail] = tail_lookup.get(tail, 0) + stock
        tail_to_gds.setdefault(tail, gds)

    if art:
        loose_entries.append(
            {
                "norm_article": art,
                "gds": gds,
                "stock": stock,
                "raw_article": str(raw_art).strip(),
            }
        )


def resolve_stock_by_etm_code(etm_code, gds_lookup):
    normalized_code = normalize_etm_code(etm_code)
    if not normalized_code:
        return 0
    return gds_lookup.get(normalized_code, 0)


def resolve_stock(row_value, lookup, tail_lookup):
    for candidate in build_match_candidates(row_value):
        stock = lookup.get(candidate, 0)
        if stock > 0:
            return stock

        if not should_use_tail_match(candidate):
            continue

        tail_key = extract_tail_key(candidate)
        if tail_key and tail_key in tail_lookup:
            return tail_lookup[tail_key]

    return 0


def resolve_gds_code(row_value, article_to_gds, tail_to_gds):
    for candidate in build_match_candidates(row_value):
        gds_code = article_to_gds.get(candidate, "")
        if gds_code:
            return gds_code

        if not should_use_tail_match(candidate):
            continue

        tail_key = extract_tail_key(candidate)
        if tail_key:
            gds_code = tail_to_gds.get(tail_key, "")
            if gds_code:
                return gds_code

    return ""


def is_safe_bulk_equivalent(candidate, norm_article):
    if not candidate or not norm_article:
        return False
    if candidate == norm_article:
        return True
    if norm_article.startswith(candidate):
        suffix = norm_article[len(candidate):]
        if not suffix:
            return False
        if re.fullmatch(r"[A-Z]", suffix):
            return True
        if re.search(r"[A-Z]", candidate) and re.fullmatch(r"0+", suffix):
            return True
    return False


def resolve_stock_loose(row_value, loose_entries):
    candidates = build_match_candidates(row_value)
    if not candidates:
        return 0

    for candidate in candidates:
        total_stock = 0
        seen = set()
        for entry in loose_entries:
            norm_article = entry["norm_article"]
            entry_key = (entry["gds"], norm_article, entry["stock"])
            if entry_key in seen:
                continue
            if is_safe_bulk_equivalent(candidate, norm_article):
                seen.add(entry_key)
                total_stock += entry["stock"]
        if total_stock > 0:
            return total_stock

    return 0


def resolve_gds_code_loose(row_value, loose_entries):
    candidates = build_match_candidates(row_value)
    if not candidates:
        return ""

    for candidate in candidates:
        for entry in loose_entries:
            if is_safe_bulk_equivalent(candidate, entry["norm_article"]) and entry["gds"]:
                return entry["gds"]

    return ""


def parse_int_stock(value):
    if value is None:
        return 0
    text = str(value).strip()
    if not text:
        return 0
    text = text.replace("\xa0", "").replace(" ", "").replace(",", ".")
    match = re.search(r"-?\d+(?:\.\d+)?", text)
    if not match:
        return 0
    try:
        return max(int(float(match.group(0))), 0)
    except ValueError:
        return 0


def pick_value(row: Dict[str, object], candidates: set) -> str:
    for key, value in row.items():
        if normalize_field_name(key) in candidates and value not in (None, ""):
            return str(value).strip()
    return ""


def row_to_stock_record(row: Dict[str, object], source: str) -> Optional[StockRecord]:
    stock_value = None
    for key, value in row.items():
        if normalize_field_name(key) in STOCK_FIELD_NAMES:
            stock_value = value
            break

    stock = parse_int_stock(stock_value)
    if stock <= 0:
        return None

    gds_code = normalize_etm_code(pick_value(row, CODE_FIELD_NAMES))
    article = pick_value(row, ARTICLE_FIELD_NAMES)
    manufacturer = pick_value(row, MANUFACTURER_FIELD_NAMES)

    if not gds_code and not article:
        return None

    return StockRecord(
        gds_code=gds_code,
        article=str(article or "").strip(),
        stock=stock,
        source=source,
        manufacturer=str(manufacturer or "").strip(),
    )


def decode_text(content: bytes) -> str:
    for encoding in ("utf-8-sig", "utf-8", "cp1251", "windows-1251", "koi8-r"):
        try:
            return content.decode(encoding)
        except UnicodeDecodeError:
            continue
    return content.decode("utf-8", errors="replace")


def parse_delimited(content: bytes, source: str) -> List[StockRecord]:
    text = decode_text(content)
    sample = text[:4096]
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=";\t,|")
    except csv.Error:
        dialect = csv.excel
        dialect.delimiter = ";"

    stream = io.StringIO(text)
    reader = csv.DictReader(stream, dialect=dialect)
    if not reader.fieldnames:
        logging.warning("No header row found in %s", source)
        return []

    header_names = [str(name or "").strip() for name in reader.fieldnames]
    normalized_headers = [normalize_field_name(name) for name in header_names]
    code_col = next(
        (idx + 1 for idx, name in enumerate(normalized_headers) if name in CODE_FIELD_NAMES),
        None,
    )
    stock_col = next(
        (idx + 1 for idx, name in enumerate(normalized_headers) if name in STOCK_FIELD_NAMES),
        None,
    )
    logging.info(
        "Delimited file %s: delimiter=%r headers=%s",
        source,
        dialect.delimiter,
        header_names[:8],
    )
    logging.info(
        "Delimited file %s: ETM code column=%s, quantity column=%s",
        source,
        code_col,
        stock_col,
    )

    records = []
    for row in reader:
        record = row_to_stock_record(row, source)
        if record:
            records.append(record)
    return records


def parse_json_payload(content: bytes, source: str) -> List[StockRecord]:
    payload = json.loads(decode_text(content))
    rows = []

    def collect(value):
        if isinstance(value, list):
            for item in value:
                collect(item)
        elif isinstance(value, dict):
            record = row_to_stock_record(value, source)
            if record:
                rows.append(record)
            for child in value.values():
                if isinstance(child, (list, dict)):
                    collect(child)

    collect(payload)
    return rows


def strip_xml_namespace(tag):
    return tag.rsplit("}", 1)[-1] if "}" in tag else tag


def parse_xml_payload(content: bytes, source: str) -> List[StockRecord]:
    root = ET.fromstring(content)
    records = []

    for elem in root.iter():
        row = {}
        for key, value in elem.attrib.items():
            row[strip_xml_namespace(key)] = value
        for child in list(elem):
            if len(list(child)) == 0:
                row[strip_xml_namespace(child.tag)] = child.text

        record = row_to_stock_record(row, source)
        if record:
            records.append(record)

    return records


def parse_xlsx_payload(content: bytes, source: str) -> List[StockRecord]:
    try:
        from openpyxl import load_workbook
    except ImportError:
        logging.warning("openpyxl is not installed; skipping XLSX file %s", source)
        return []

    workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    records = []
    for sheet in workbook.worksheets:
        rows = sheet.iter_rows(values_only=True)
        header = next(rows, None)
        if not header:
            continue
        header_names = [str(cell or "").strip() for cell in header]
        if not any(normalize_field_name(name) in STOCK_FIELD_NAMES for name in header_names):
            continue
        for values in rows:
            row = dict(zip(header_names, values))
            record = row_to_stock_record(row, f"{source}:{sheet.title}")
            if record:
                records.append(record)
    workbook.close()
    return records


def parse_zip_payload(content: bytes, source: str) -> List[StockRecord]:
    records = []
    with zipfile.ZipFile(io.BytesIO(content)) as archive:
        for name in archive.namelist():
            if name.endswith("/"):
                continue
            with archive.open(name) as fh:
                nested_content = fh.read()
            records.extend(parse_stock_payload(nested_content, f"{source}!{name}"))
    return records


def parse_stock_payload(content: bytes, source: str) -> List[StockRecord]:
    suffix = Path(source.split("!", 1)[-1]).suffix.lower()
    try:
        if suffix == ".zip":
            return parse_zip_payload(content, source)
        if suffix in {".xlsx", ".xlsm"}:
            return parse_xlsx_payload(content, source)
        if suffix == ".json":
            return parse_json_payload(content, source)
        if suffix == ".xml":
            return parse_xml_payload(content, source)
        if suffix in {".csv", ".txt", ".tsv", ".dat"}:
            return parse_delimited(content, source)

        stripped = content.lstrip()
        if stripped.startswith(b"PK"):
            return parse_zip_payload(content, source)
        if stripped.startswith(b"{") or stripped.startswith(b"["):
            return parse_json_payload(content, source)
        if stripped.startswith(b"<"):
            return parse_xml_payload(content, source)
        return parse_delimited(content, source)
    except Exception as exc:
        logging.warning("Could not parse %s: %s", source, exc)
        return []


def get_ftp_password():
    if FTP_PASSWORD:
        return FTP_PASSWORD

    try:
        result = subprocess.run(
            [
                "security",
                "find-generic-password",
                "-a",
                FTP_USER,
                "-s",
                "checksheets_etm_ftp",
                "-w",
            ],
            check=False,
            capture_output=True,
            text=True,
            timeout=10,
        )
    except Exception as exc:
        logging.warning("Could not query macOS Keychain for ETM FTP password: %s", exc)
        return ""

    if result.returncode != 0:
        return ""

    return result.stdout.strip()


def connect_ftp():
    password = get_ftp_password()
    if not FTP_USER or not password:
        raise RuntimeError(
            "Set ETM_FTP_USER and ETM_FTP_PASSWORD, or add password to macOS Keychain "
            "service checksheets_etm_ftp before running FTP stock sync"
        )

    if FTP_TLS_MODE not in {"auto", "require", "disable"}:
        raise ValueError("ETM_FTP_TLS must be auto, require, or disable")

    if FTP_TLS_MODE != "disable":
        try:
            ftp_tls = FTP_TLS()
            ftp_tls.connect(FTP_HOST, FTP_PORT, timeout=FTP_TIMEOUT)
            ftp_tls.login(FTP_USER, password)
            ftp_tls.prot_p()
            ftp_tls.set_pasv(True)
            logging.info("Connected to FTPS %s:%s as %s", FTP_HOST, FTP_PORT, FTP_USER)
            return ftp_tls
        except Exception as exc:
            if FTP_TLS_MODE == "require":
                raise
            logging.warning(
                "FTPS connection failed (%s); falling back to plain FTP for %s:%s",
                exc,
                FTP_HOST,
                FTP_PORT,
            )

    ftp = FTP()
    ftp.connect(FTP_HOST, FTP_PORT, timeout=FTP_TIMEOUT)
    ftp.login(FTP_USER, password)
    ftp.set_pasv(True)
    logging.warning("Connected to plain FTP %s:%s as %s", FTP_HOST, FTP_PORT, FTP_USER)
    return ftp


def is_ftp_dir(ftp, remote_path):
    current = ftp.pwd()
    try:
        ftp.cwd(remote_path)
        ftp.cwd(current)
        return True
    except Exception:
        try:
            ftp.cwd(current)
        except Exception:
            pass
        return False


def mdtm(ftp, remote_path):
    try:
        response = ftp.sendcmd(f"MDTM {remote_path}")
    except Exception:
        return None
    match = re.search(r"(\d{14})", response)
    return match.group(1) if match else None


def walk_ftp_files(ftp, remote_dir) -> List[FtpFile]:
    remote_dir = "/" + remote_dir.strip("/")
    files = []

    try:
        for name, facts in ftp.mlsd(remote_dir):
            if name in {".", ".."}:
                continue
            remote_path = posixpath.join(remote_dir, name)
            if facts.get("type") == "dir":
                files.extend(walk_ftp_files(ftp, remote_path))
            elif facts.get("type") == "file":
                size = int(facts["size"]) if str(facts.get("size", "")).isdigit() else None
                files.append(
                    FtpFile(
                        remote_path=remote_path,
                        size=size,
                        modified=facts.get("modify") or mdtm(ftp, remote_path),
                    )
                )
        return files
    except Exception:
        pass

    try:
        entries = ftp.nlst(remote_dir)
    except error_perm as exc:
        logging.warning("Could not list %s: %s", remote_dir, exc)
        return []

    for entry in entries:
        name = entry.rstrip("/").split("/")[-1]
        if name in {".", ".."}:
            continue
        remote_path = entry if entry.startswith("/") else posixpath.join(remote_dir, entry)
        if is_ftp_dir(ftp, remote_path):
            files.extend(walk_ftp_files(ftp, remote_path))
        else:
            files.append(FtpFile(remote_path=remote_path, modified=mdtm(ftp, remote_path)))

    return files


def choose_files_for_processing(files: List[FtpFile], mode: str) -> List[FtpFile]:
    if mode == "all":
        return sorted(files, key=lambda item: item.remote_path)
    if not files:
        return []

    latest = max(files, key=lambda item: (item.modified or "", item.remote_path))
    return [latest]


def ftp_modify_date(modified):
    if not modified or not re.fullmatch(r"\d{14}", str(modified)):
        return None
    return str(modified)[:8]


def today_ftp_date():
    return datetime.now(timezone.utc).strftime("%Y%m%d")


def yesterday_ftp_date():
    return (datetime.now(timezone.utc) - timedelta(days=1)).strftime("%Y%m%d")


def filter_today_files(files: List[FtpFile]) -> List[FtpFile]:
    if not FTP_REQUIRE_TODAY:
        return files

    today = today_ftp_date()
    today_files = [item for item in files if ftp_modify_date(item.modified) == today]
    if today_files:
        return today_files

    yesterday = yesterday_ftp_date()
    yesterday_files = [
        item for item in files if ftp_modify_date(item.modified) == yesterday
    ]
    if yesterday_files:
        logging.warning(
            "No FTP files dated %s; using %s yesterday fallback file(s)",
            today,
            len(yesterday_files),
        )
        return yesterday_files

    skipped = len(files)
    if skipped:
        logging.info(
            "Skipped %s FTP files: no files dated %s or fallback date %s",
            skipped,
            today,
            yesterday,
        )
    return []


def file_fingerprint(ftp_file: FtpFile):
    return {
        "remote_path": ftp_file.remote_path,
        "size": ftp_file.size,
        "modified": ftp_file.modified,
    }


def load_ftp_state():
    try:
        return json.loads(FTP_STATE_PATH.read_text(encoding="utf-8"))
    except FileNotFoundError:
        return {}
    except Exception as exc:
        logging.warning("Could not read FTP state %s: %s", FTP_STATE_PATH, exc)
        return {}


def save_ftp_state(state):
    FTP_STATE_PATH.parent.mkdir(parents=True, exist_ok=True)
    FTP_STATE_PATH.write_text(
        json.dumps(state, ensure_ascii=False, indent=2, sort_keys=True),
        encoding="utf-8",
    )


def local_path_for(remote_path):
    return FTP_LOCAL_ROOT / remote_path.lstrip("/")


def download_ftp_file(ftp, remote_path) -> bytes:
    started_at = time.monotonic()
    chunks = []
    bytes_received = 0

    def collect(chunk):
        nonlocal bytes_received
        bytes_received += len(chunk)
        chunks.append(chunk)

    try:
        ftp.retrbinary(f"RETR {remote_path}", collect)
    except Exception as exc:
        logging.warning(
            "FTP download failed: path=%s bytes=%s duration=%.2fs error=%s(%s)",
            remote_path,
            bytes_received,
            time.monotonic() - started_at,
            type(exc).__name__,
            str(exc) or "<empty>",
        )
        raise

    content = b"".join(chunks)
    logging.info(
        "FTP download complete: path=%s bytes=%s duration=%.2fs",
        remote_path,
        len(content),
        time.monotonic() - started_at,
    )
    return content


def validate_ftp_csv(content: bytes, source: str) -> Dict[str, object]:
    """Проверить весь CSV до сохранения и обработки источника."""
    text = decode_text(content)
    if not text.strip():
        raise ValueError(f"{source}: empty CSV")

    sample = text[:4096]
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=";\t,|")
    except csv.Error:
        dialect = csv.excel
        dialect.delimiter = ";"

    reader = csv.reader(io.StringIO(text), dialect=dialect, strict=True)
    fieldnames = next(reader, None)
    if not fieldnames:
        raise ValueError(f"{source}: missing CSV header")

    normalized_headers = {normalize_field_name(name) for name in fieldnames if name is not None}
    if not normalized_headers.intersection(CODE_FIELD_NAMES):
        raise ValueError(f"{source}: missing ETM code column")
    if not normalized_headers.intersection(STOCK_FIELD_NAMES):
        raise ValueError(f"{source}: missing stock quantity column")

    data_rows = 0
    for _row in reader:
        data_rows += 1
    if data_rows == 0:
        raise ValueError(f"{source}: CSV has no data rows")

    return {
        "header_columns": len(fieldnames),
        "data_rows": data_rows,
        "has_data": True,
    }


def save_ftp_cache(remote_path: str, content: bytes) -> None:
    """Атомарно сохранить только полный и проверенный FTP-файл."""
    local_path = local_path_for(remote_path)
    local_path.parent.mkdir(parents=True, exist_ok=True)
    temporary = local_path.with_name(f".{local_path.name}.tmp")
    temporary.write_bytes(content)
    temporary.replace(local_path)


def build_lookup_from_records(records: Iterable[StockRecord]):
    lookup = {}
    gds_lookup = {}
    tail_lookup = {}
    article_to_gds = {}
    tail_to_gds = {}
    loose_entries = []

    count = 0
    for record in records:
        count += 1
        add_stock_to_lookup(
            lookup,
            gds_lookup,
            tail_lookup,
            article_to_gds,
            tail_to_gds,
            loose_entries,
            {
                "GdsCode": record.gds_code,
                "Article": record.article,
                "RemInfo": record.stock,
            },
        )
    return {
        "records": count,
        "lookup": lookup,
        "gds_lookup": gds_lookup,
        "tail_lookup": tail_lookup,
        "article_to_gds": article_to_gds,
        "tail_to_gds": tail_to_gds,
        "loose_entries": loose_entries,
    }


def summarize_stock_records(records: List[StockRecord], label: str):
    if not records:
        logging.info("%s parsed stock summary: records=0", label)
        return

    total_qty = sum(record.stock for record in records)
    unique_codes = len({record.gds_code for record in records if record.gds_code})
    samples = [
        f"{record.gds_code}:{record.stock}"
        for record in records[:5]
    ]
    logging.info(
        "%s parsed stock summary: records=%s unique_etm_codes=%s total_qty=%s samples=%s",
        label,
        len(records),
        unique_codes,
        total_qty,
        samples,
    )


def fetch_warehouse_stock_lookup(ftp, remote_dir, label, process_mode, state_key, state, force=False):
    files = filter_today_files(walk_ftp_files(ftp, remote_dir))
    selected_files = choose_files_for_processing(files, process_mode)
    last_processed = state.get(state_key, {}).get("files", [])
    selected_fingerprints = [file_fingerprint(item) for item in selected_files]
    already_processed = (
        bool(selected_files) and selected_fingerprints == last_processed and not force
    )

    logging.info(
        "%s FTP files: found=%s selected=%s mode=%s force=%s already_processed=%s dir=%s",
        label,
        len(files),
        len(selected_files),
        process_mode,
        force,
        already_processed,
        remote_dir,
    )

    if already_processed:
        return build_lookup_from_records([]), selected_fingerprints, False

    records = []
    for ftp_file in selected_files:
        logging.info(
            "%s download: %s size=%s modified=%s",
            label,
            ftp_file.remote_path,
            ftp_file.size,
            ftp_file.modified,
        )
        content = download_ftp_file(ftp, ftp_file.remote_path)
        suffix = Path(ftp_file.remote_path).suffix.lower()
        if suffix in {".csv", ".txt", ".tsv", ".dat"}:
            validation = validate_ftp_csv(content, ftp_file.remote_path)
            logging.info(
                "%s CSV validated: path=%s header_columns=%s data_rows=%s has_data=%s",
                label,
                ftp_file.remote_path,
                validation["header_columns"],
                validation["data_rows"],
                validation["has_data"],
            )
        elif not content.strip():
            raise ValueError(f"{ftp_file.remote_path}: empty FTP payload")
        parsed = parse_stock_payload(content, ftp_file.remote_path)
        records.extend(parsed)
        # Не заменяем локальный файл до успешного полного скачивания,
        # структурной проверки CSV и завершения парсинга.
        save_ftp_cache(ftp_file.remote_path, content)
        logging.info("%s parsed %s stock records from %s", label, len(parsed), ftp_file.remote_path)

    summarize_stock_records(records, label)
    lookup_bundle = build_lookup_from_records(records)
    logging.info(
        "%s stock lookup: records=%s gds=%s articles=%s",
        label,
        lookup_bundle["records"],
        len(lookup_bundle["gds_lookup"]),
        len(lookup_bundle["lookup"]),
    )
    return lookup_bundle, selected_fingerprints, bool(selected_files)


def close_ftp(ftp):
    if ftp is None:
        return
    try:
        ftp.quit()
    except Exception:
        try:
            ftp.close()
        except Exception:
            pass


def fetch_warehouse_stock_lookup_with_retry(
    remote_dir, label, process_mode, state_key, state, force=False
):
    if FTP_WAREHOUSE_RETRIES < 1:
        raise ValueError("ETM_FTP_WAREHOUSE_RETRIES must be at least 1")

    last_exc = None
    for attempt in range(1, FTP_WAREHOUSE_RETRIES + 1):
        ftp = None
        try:
            ftp = connect_ftp()
            return fetch_warehouse_stock_lookup(
                ftp, remote_dir, label, process_mode, state_key, state, force
            )
        except (EOFError, OSError, socket.timeout) as exc:
            last_exc = exc
            if attempt >= FTP_WAREHOUSE_RETRIES:
                break
            delay = FTP_WAREHOUSE_RETRY_DELAY * attempt
            logging.warning(
                "%s FTP read failed on attempt %s/%s: %s. Reconnecting in %.1fs",
                label,
                attempt,
                FTP_WAREHOUSE_RETRIES,
                exc,
                delay,
            )
            time.sleep(delay)
        finally:
            close_ftp(ftp)

    if last_exc is None:
        raise RuntimeError(f"{label} FTP read failed without a captured exception")
    raise last_exc


def update_sheet_range_with_retry(ws, range_name, values):
    last_exc = None
    for attempt in range(1, SHEETS_UPDATE_RETRIES + 1):
        try:
            ws.update(range_name=range_name, values=values)
            return
        except (
            requests.exceptions.SSLError,
            requests.exceptions.ConnectionError,
            requests.exceptions.Timeout,
        ) as exc:
            last_exc = exc
            if attempt >= SHEETS_UPDATE_RETRIES:
                break
            delay = SHEETS_UPDATE_RETRY_DELAY * attempt
            logging.warning(
                "Google Sheets update failed for %s on attempt %s/%s: %s. Retrying in %.1fs",
                range_name,
                attempt,
                SHEETS_UPDATE_RETRIES,
                exc,
                delay,
            )
            time.sleep(delay)

    raise last_exc


def compute_sheet_values(all_data, etm_code_col, nsb_bundle, smr_bundle):
    nsb_results = []
    smr_results = []
    matched_nsb = 0
    matched_smr = 0
    missing_etm_codes = 0

    for row in all_data[1:]:
        etm_code = get_sheet_etm_code(row, etm_code_col)
        if not etm_code:
            missing_etm_codes += 1

        stock_nsb = resolve_stock_by_etm_code(etm_code, nsb_bundle["gds_lookup"])
        stock_smr = resolve_stock_by_etm_code(etm_code, smr_bundle["gds_lookup"])

        if stock_nsb > 0:
            matched_nsb += 1
        if stock_smr > 0:
            matched_smr += 1

        nsb_results.append([stock_nsb])
        smr_results.append([stock_smr])

    return {
        "nsb_results": nsb_results,
        "smr_results": smr_results,
        "matched_nsb": matched_nsb,
        "matched_smr": matched_smr,
        "missing_etm_codes": missing_etm_codes,
    }


def compute_single_stock_column(all_data, etm_code_col, stock_lookup):
    results = []
    matched = 0
    missing_etm_codes = 0

    for row in all_data[1:]:
        etm_code = get_sheet_etm_code(row, etm_code_col)
        if not etm_code:
            missing_etm_codes += 1

        stock = resolve_stock_by_etm_code(etm_code, stock_lookup)
        if stock > 0:
            matched += 1
        results.append([stock])

    return {
        "results": results,
        "matched": matched,
        "missing_etm_codes": missing_etm_codes,
    }


def compute_feron_mapping(all_data, mapping, match_col):
    code_results = []
    stock_codes = []
    matched = 0

    for row in all_data[1:]:
        match_key = normalize_mapping_key(
            row[match_col - 1] if len(row) >= match_col else ""
        )
        etm_code = mapping.get(match_key, "")
        if etm_code:
            matched += 1
        code_results.append([etm_code])
        stock_codes.append(etm_code)

    return code_results, stock_codes, matched


def compute_stock_values_from_codes(etm_codes, stock_lookup):
    results = []
    matched = 0
    for etm_code in etm_codes:
        stock = resolve_stock_by_etm_code(etm_code, stock_lookup)
        if stock > 0:
            matched += 1
        results.append([stock])
    return {"results": results, "matched": matched}


def compute_feron_stock_by_model_brand(all_data, stock_lookup, model_col, brand_col):
    results = []
    matched = 0
    missing_keys = 0

    for row in all_data[1:]:
        model = normalize(row[model_col - 1] if len(row) >= model_col else "")
        brand = normalize(row[brand_col - 1] if len(row) >= brand_col else "")
        if not model or not brand:
            missing_keys += 1
        stock = stock_lookup.get((model, brand), 0)
        if stock > 0:
            matched += 1
        results.append([stock])

    return {
        "results": results,
        "matched": matched,
        "missing_keys": missing_keys,
    }


def sync(process_mode=FTP_PROCESS_MODE, dry_run=False, force=False):
    logging.info("--- STARTING ETM FTP STOCK SYNC ---")
    start_time = time.time()
    process_mode = (process_mode or "latest").strip().lower()
    if process_mode not in {"latest", "all"}:
        raise ValueError("ETM_FTP_PROCESS_MODE must be 'latest' or 'all'")

    logging.info(
        "ETM FTP settings: host=%s port=%s user=%s tls=%s mode=%s require_today=%s force=%s",
        FTP_HOST,
        FTP_PORT,
        FTP_USER,
        FTP_TLS_MODE,
        process_mode,
        FTP_REQUIRE_TODAY,
        force,
    )
    logging.info(
        "StreamSupps matching: header 'CODES' -> FTP columns A 'Код ЭТМ' and D 'Количество'",
    )
    logging.info(
        "Warehouse mapping: %s => %s, %s => %s",
        WAREHOUSE_DIRS["smr"]["remote_dir"],
        WAREHOUSE_DIRS["smr"]["header"],
        WAREHOUSE_DIRS["nsb"]["remote_dir"],
        WAREHOUSE_DIRS["nsb"]["header"],
    )

    state = load_ftp_state()
    smr_bundle, smr_files, smr_has_new_files = fetch_warehouse_stock_lookup_with_retry(
        WAREHOUSE_DIRS["smr"]["remote_dir"],
        WAREHOUSE_DIRS["smr"]["label"],
        process_mode,
        "smr",
        state,
        force,
    )
    nsb_bundle, nsb_files, nsb_has_new_files = fetch_warehouse_stock_lookup_with_retry(
        WAREHOUSE_DIRS["nsb"]["remote_dir"],
        WAREHOUSE_DIRS["nsb"]["label"],
        process_mode,
        "nsb",
        state,
        force,
    )

    if not smr_has_new_files and not nsb_has_new_files:
        logging.info("No new ETM FTP files to process; Google Sheets values were not written")
        return 0

    if smr_has_new_files and smr_bundle["records"] <= 0:
        logging.warning("Samara FTP file was found but no stock records were parsed")
    if nsb_has_new_files and nsb_bundle["records"] <= 0:
        logging.warning("Novosibirsk FTP file was found but no stock records were parsed")

    if smr_bundle["records"] <= 0 and nsb_bundle["records"] <= 0:
        logging.info("No parsed ETM stock records; Google Sheets values were not written")
        return 0

    ws = gsheets_utils.get_worksheet(STREAM_SUPPS_SHEET_NAME)
    all_data = ws.get_all_values()
    if not all_data:
        raise RuntimeError(f"{STREAM_SUPPS_SHEET_NAME} sheet is empty")

    etm_columns = gsheets_utils.resolve_header_columns(
        all_data[0], ETM_TR_SCHEMA, STREAM_SUPPS_SHEET_NAME
    )
    col_stock_nsb = etm_columns["stock_nsb"]
    col_stock_smr = etm_columns["stock_smr"]

    computed = compute_sheet_values(
        all_data, etm_columns["etm_code"], nsb_bundle, smr_bundle
    )

    logging.info(
        "MATCHED NSB: %s / %s",
        computed["matched_nsb"],
        len(computed["nsb_results"]),
    )
    logging.info(
        "MATCHED SMR: %s / %s",
        computed["matched_smr"],
        len(computed["smr_results"]),
    )
    logging.info("StreamSupps rows without 'CODES': %s", computed["missing_etm_codes"])

    if dry_run:
        logging.info("Dry run enabled; Google Sheets values were not written")
        return 0

    logging.info("Writing StreamSupps ETM values to Google Sheets...")

    wrote_any = False
    if nsb_bundle["records"] > 0:
        nsb_range = (
            f"{gspread.utils.rowcol_to_a1(2, col_stock_nsb)}:"
            f"{gspread.utils.rowcol_to_a1(1 + len(computed['nsb_results']), col_stock_nsb)}"
        )
        update_sheet_range_with_retry(ws, nsb_range, computed["nsb_results"])
        state["nsb"] = {
            "files": nsb_files,
            "processed_at": datetime.now(timezone.utc).isoformat(),
        }
        wrote_any = True
    else:
        logging.info("Novosibirsk values were not written because no records were parsed")

    if smr_bundle["records"] > 0:
        smr_range = (
            f"{gspread.utils.rowcol_to_a1(2, col_stock_smr)}:"
            f"{gspread.utils.rowcol_to_a1(1 + len(computed['smr_results']), col_stock_smr)}"
        )
        update_sheet_range_with_retry(ws, smr_range, computed["smr_results"])
        state["smr"] = {
            "files": smr_files,
            "processed_at": datetime.now(timezone.utc).isoformat(),
        }
        wrote_any = True
    else:
        logging.info("Samara values were not written because no records were parsed")

    if not wrote_any:
        logging.info("No StreamSupps ETM ranges were written")
        return 0

    save_ftp_state(state)
    logging.info("StreamSupps ETM values written successfully")
    logging.info("SUCCESS! Time: %.1fs", time.time() - start_time)
    return 0


def parse_args():
    parser = argparse.ArgumentParser(description="Sync ETM TR stocks from ETM FTPS export")
    parser.add_argument(
        "--mode",
        choices=("latest", "all"),
        default=FTP_PROCESS_MODE,
        help="Process only latest file per warehouse directory or all files",
    )
    parser.add_argument(
        "--dry-run",
        action="store_true",
        help="Download and parse FTP files but do not write Google Sheets",
    )
    parser.add_argument(
        "--force",
        action="store_true",
        help="Process the selected FTP files even if they were already processed",
    )
    return parser.parse_args()


if __name__ == "__main__":
    args = parse_args()
    try:
        sys.exit(sync(process_mode=args.mode, dry_run=args.dry_run, force=args.force))
    except Exception as exc:
        logging.exception("CRITICAL ERROR: %s", exc)
        sys.exit(1)
